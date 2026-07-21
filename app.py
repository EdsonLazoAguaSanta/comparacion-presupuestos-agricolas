import streamlit as st
import pandas as pd
import numpy as np
try:
    import pyodbc
    _PYODBC_OK = True
except ImportError:
    _PYODBC_OK = False
import requests
import hashlib, hmac
import time
import base64
import socket
from io import BytesIO
from openpyxl import load_workbook

# ═══════════════════════════════════════════════════════════════════════════
#  Comparación de Presupuestos Agrícolas · Costo x Hectárea por Especie/Fundo
#  App independiente ("acceso aparte") — no depende del Informe de Cierre.
#
#  ALCANCE ACTUAL (2026-07-21): SOLO PRESUPUESTOS, temporadas 2024 en
#  adelante.
#
#  Vistas:
#   · Resumen (por defecto): 5 categorías de costo (Costo Mano de obra,
#     Costo MO Cosecha, Fitosanitarios, Generales, Depreciación).
#   · Detalle (OCULTA — se accede agregando ?view=detalle a la URL, sin link
#     visible todavía): una fila por TipoDetG del presupuesto, y la Mano de
#     Obra desglosada por Faena (columna N_Faena).
#
#  METODOLOGÍA / SUPUESTOS (revisar con datos reales antes de confiar 100%):
#   · Montos de presupuesto (todo menos Mano de Obra): tabla
#     [Presupuesto].[dbo].[Presupuesto], filtrada por Temporada (mismo
#     criterio que usa el PBI de Gastos).
#   · Costo Mano de obra / Costo MO Cosecha: tabla
#     [Presupuesto].[dbo].[ManoDeObra] (mismo servidor SQL), que tiene su
#     propia columna N_Faena por Ccosto — las filas cuyo N_Faena contiene
#     "Cosecha" van a "Costo MO Cosecha", el resto a "Costo Mano de obra".
#   · Hectáreas/Producción por Ccosto: SUM total del Ccosto en
#     [CRUZDELSUR2].[Produccion].[softland].[ProduccionFundos] para esa
#     Temporada (no se separa por variedad).
#   · Especie/Variedad por Ccosto: NO se usa pf.Especie/pf.Variedad de
#     ProduccionFundos porque contiene códigos internos poco confiables
#     para mostrar al usuario (ej. "B062"). En su lugar se usa el mismo
#     criterio que ya aplica la pestaña "EERR Proyectado" del Informe de
#     Cierre: la columna Variedad2 (o Variedad si no existe) de la hoja
#     "Captura Ingresos" del archivo de Ingresos, indexada por Ccosto. Ese
#     Excel solo existe para la temporada actual, así que este mapeo
#     Ccosto→Especie/Variedad se usa como maestro FIJO para TODAS las
#     temporadas de presupuesto.
#   · "Generales" (vista Resumen) agrupa TipoDetG: Otros Gastos, Otros
#     Personal, Petróleo, Comercial CdS, Corrección monetaria, Gastos
#     financieros.
#   · El costo de un Ccosto se asigna completo a cada Especie/Variedad que
#     ahí se cultive (no se prorratea) — mismo criterio ya usado en el
#     resto del sistema (EERR Proyectado).
# ═══════════════════════════════════════════════════════════════════════════

# ── Secretos: SIEMPRE desde st.secrets, nunca hardcodeados en el código
#    (este archivo se sube a un repo Git para desplegar en Streamlit Cloud).
#    Ver secrets.toml.example para el formato esperado. Se aceptan varios
#    nombres de clave por compatibilidad con el secrets.toml ya existente
#    del Informe de Cierre (GRAPH_TENANT_ID, etc.). ───────────────────────────
def _req_secret(paths, label):
    """paths: lista de rutas candidatas; cada una string (raíz) o tupla (sección, clave)."""
    for path in paths:
        try:
            node = st.secrets
            for k in (path if isinstance(path, tuple) else (path,)):
                node = node[k]
            if node:
                return node
        except Exception:
            continue
    st.error(
        f"⚠️ Falta configurar `st.secrets` para **{label}**.\n\n"
        f"En local: copia `secrets.toml.example` a `.streamlit/secrets.toml` y complétalo.\n"
        f"En Streamlit Cloud: agrégalo en Settings → Secrets de la app.")
    st.stop()


# ── Config SharePoint (mismo tenant/sitio que el Informe de Cierre) ─────────
TENANT_ID     = _req_secret([("azure", "tenant_id"), "GRAPH_TENANT_ID", "TENANT_ID"], "tenant_id de Azure")
CLIENT_ID     = _req_secret([("azure", "client_id"), "GRAPH_CLIENT_ID", "CLIENT_ID"], "client_id de Azure")
CLIENT_SECRET = _req_secret([("azure", "client_secret"), "GRAPH_CLIENT_SECRET", "CLIENT_SECRET"], "client_secret de Azure")
SITE_NAME     = st.secrets.get("azure", {}).get("site_name", "Gestion")
FILE_NAME     = "Ingresos 2025-2026 para PBI - copia.xlsx"   # maestro Ccosto→Especie/Variedad
SHEET_CANDS   = ["Captura Ingresos", "Hoja3", "Sheet3", "tblCapturaIngresos"]
_ANCHOR_FILES = ["GASTOS_NT.xlsx", "APP_WEB_NO_MOVER_NI_TOCAR.xlsx", FILE_NAME]
PASSWORDS_FILE = "passwords.json"

SECRET_KEY = _req_secret(["app_secret_key", "APP_SECRET_KEY"], "app_secret_key")
# ── Usuarios con acceso a ESTA app (lista propia, independiente del
#    Informe de Cierre). Añadir/quitar emails aquí. ─────────────────────────
COMPARACION_USERS = {"elazo@aguasanta.cl", "flow@aguasanta.cl"}
SESSION_TTL = 48 * 3600

FUNDOS_ORDEN = ["La Torina", "Santa Amelia", "Santa Ana", "El Cóndor", "El Carmelo",
                "Agua Santa", "Santa Norma", "Villaseca", "Malihuito", "Lisonjera"]

TEMPORADA_MIN = 24   # 2024 en adelante

TIPOS_OTROS = {"otros gastos", "otros personal", "petroleo", "comercial cds",
               "corrección monetaria", "correccion monetaria", "gastos financieros"}

# Orden de TipoDetG para la vista Detalle (Mano de obra se excluye de acá:
# se desglosa aparte por Faena desde [Presupuesto].[dbo].[ManoDeObra]).
TIPOS_DETALLE_ORDEN = ["Fitosanitarios", "Otros Gastos", "Otros Personal", "Depreciación",
                        "Petroleo", "Comercial CdS", "Corrección monetaria", "Gastos financieros"]


# ═══════════════════════════════════════════════════════════════════════════
#  AUTENTICACIÓN (mismo esquema que el Informe de Cierre: token en query
#  param 'sid', contraseñas en st.secrets["usuarios"] o passwords.json en
#  SharePoint). Restringido a COMPARACION_USERS.
# ═══════════════════════════════════════════════════════════════════════════

def _sign(user, exp):
    raw = f"{user}|{exp}|{SECRET_KEY}"
    return hmac.new(SECRET_KEY.encode(), raw.encode(), hashlib.sha256).hexdigest()[:16]


def _make_token(user, exp):
    sig = _sign(user, exp)
    raw = f"{user}|{exp}|{sig}"
    return base64.urlsafe_b64encode(raw.encode()).decode()


def _parse_token(token_b64):
    try:
        raw = base64.urlsafe_b64decode(token_b64.encode()).decode()
        parts = raw.split("|")
        if len(parts) != 3:
            return None
        user, exp_s, sig = parts
        if float(exp_s) < time.time():
            return None
        if not hmac.compare_digest(sig, _sign(user, exp_s)):
            return None
        if user not in COMPARACION_USERS:
            return None
        return user
    except Exception:
        return None


@st.cache_data(ttl=3500)
def get_token():
    r = requests.post(
        f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token",
        data={"grant_type": "client_credentials", "client_id": CLIENT_ID,
              "client_secret": CLIENT_SECRET, "scope": "https://graph.microsoft.com/.default"})
    r.raise_for_status()
    return r.json()["access_token"]


@st.cache_data(ttl=3500)
def get_site_id(token):
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/servicioscruzdelsur.sharepoint.com:/sites/{SITE_NAME}",
        headers={"Authorization": f"Bearer {token}"})
    r.raise_for_status()
    return r.json()["id"]


def _drive_search(token, site_id, q):
    headers = {"Authorization": f"Bearer {token}"}
    out = []
    try:
        r = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives", headers=headers)
        r.raise_for_status()
        for drive in r.json().get("value", []):
            did = drive["id"]
            try:
                r2 = requests.get(f"https://graph.microsoft.com/v1.0/drives/{did}/root/search(q='{q}')",
                                   headers=headers)
                if r2.status_code == 200:
                    for it in r2.json().get("value", []):
                        out.append((did, it))
            except Exception:
                continue
    except Exception:
        pass
    return out


def get_file_location(token, site_id, filename):
    headers = {"Authorization": f"Bearer {token}"}
    want = filename.strip().lower()
    base = want.rsplit(".", 1)[0].replace(" ", "")

    for did, it in _drive_search(token, site_id, filename):
        if it.get("name", "").strip().lower() == want:
            return did, it["id"]

    for anchor in _ANCHOR_FILES:
        for did, it in _drive_search(token, site_id, anchor):
            if it.get("name", "").strip().lower() != anchor.strip().lower():
                continue
            folder = (it.get("parentReference", {}) or {}).get("id")
            if not folder:
                try:
                    ri = requests.get(f"https://graph.microsoft.com/v1.0/drives/{did}/items/{it['id']}",
                                       headers=headers)
                    folder = (ri.json().get("parentReference", {}) or {}).get("id")
                except Exception:
                    folder = None
            if not folder:
                continue
            url = (f"https://graph.microsoft.com/v1.0/drives/{did}/items/{folder}"
                   f"/children?$top=200&$select=id,name")
            while url:
                try:
                    rc = requests.get(url, headers=headers)
                    if rc.status_code != 200:
                        break
                    j = rc.json()
                except Exception:
                    break
                for ch in j.get("value", []):
                    nm = ch.get("name", "").strip().lower()
                    if nm == want or (nm.replace(" ", "").startswith(base) and nm.endswith(".xlsx")):
                        return did, ch["id"]
                url = j.get("@odata.nextLink")
            break
    return None, None


def load_passwords(token, site_id):
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/search(q='{PASSWORDS_FILE}')",
        headers={"Authorization": f"Bearer {token}"})
    if r.status_code != 200:
        return {}
    iid = None
    for item in r.json().get("value", []):
        if item.get("name", "") == PASSWORDS_FILE:
            iid = item["id"]
            break
    if iid is None:
        return {}
    try:
        rc = requests.get(
            f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{iid}/content",
            headers={"Authorization": f"Bearer {token}"})
        rc.raise_for_status()
        import json as _json
        return _json.loads(rc.content.decode("utf-8"))
    except Exception:
        return {}


def check_password(user, pwd_ingresada, token, site_id):
    passwords = load_passwords(token, site_id)
    if user in passwords:
        return passwords[user] == hashlib.sha256(pwd_ingresada.encode()).hexdigest()
    usuarios = st.secrets.get("usuarios", {})
    return usuarios.get(user, {}).get("password", "") == pwd_ingresada


def check_login():
    if st.session_state.get("logged_in"):
        elapsed = time.time() - st.session_state.get("login_time", 0)
        if elapsed < SESSION_TTL:
            return True
        for k in ["logged_in", "login_time", "user_email"]:
            st.session_state.pop(k, None)

    qp = st.query_params
    token = qp.get("sid", None)
    if token:
        user = _parse_token(token)
        if user:
            st.session_state["logged_in"] = True
            st.session_state["login_time"] = time.time()
            st.session_state["user_email"] = user
            return True
        else:
            st.query_params.clear()

    st.set_page_config(page_title="Comparación Presupuestos | Agua Santa", layout="centered", page_icon="🌿")
    st.markdown("""
<style>
#MainMenu,footer,header{visibility:hidden}
.block-container{padding-top:3rem!important}
.stApp{background:#F4F6F7}
div[data-testid="stFormSubmitButton"]>button{
    background:linear-gradient(135deg,#0C383B,#145f63)!important;color:#fff!important;
    font-weight:700!important;border:none!important;border-radius:8px!important}
</style>
""", unsafe_allow_html=True)
    st.markdown("""
<div style="text-align:center;margin-bottom:28px">
  <svg width="48" height="48" viewBox="0 0 24 24" fill="none" style="margin:0 auto 10px;display:block">
    <path d="M12 2C8 2 4 6 4 10c0 5 8 12 8 12s8-7 8-12c0-4-4-8-8-8z" fill="#2a6630" opacity=".95"/>
    <path d="M12 6v8M8 10h8" stroke="#fff" stroke-width="1.5" stroke-linecap="round"/>
  </svg>
  <div style="color:#072b0a;font-size:1.25rem;font-weight:700">Empresas Agua Santa</div>
  <div style="color:#072b0a;font-size:.8rem">Comparación de Presupuestos · Costo x Hectárea</div>
</div>
""", unsafe_allow_html=True)

    with st.form("login_form"):
        usuario = st.text_input("Usuario", placeholder="tu.nombre@aguasanta.cl")
        password = st.text_input("Contraseña", type="password")
        submitted = st.form_submit_button("Ingresar", use_container_width=True, type="primary")

    if submitted:
        u = usuario.strip().lower()
        if u not in COMPARACION_USERS:
            st.error("❌ Este usuario no tiene acceso a esta aplicación.")
        else:
            try:
                tok = get_token(); sid = get_site_id(tok)
                ok = check_password(u, password, tok, sid)
            except Exception:
                usuarios = st.secrets.get("usuarios", {})
                ok = usuarios.get(u, {}).get("password", "") == password
            if ok:
                exp = time.time() + SESSION_TTL
                token = _make_token(u, str(exp))
                st.session_state["logged_in"] = True
                st.session_state["login_time"] = time.time()
                st.session_state["user_email"] = u
                st.query_params["sid"] = token
                st.rerun()
            else:
                st.error("❌ Usuario o contraseña incorrectos.")
    return False


if not check_login():
    st.stop()

user_email = st.session_state["user_email"]


# ═══════════════════════════════════════════════════════════════════════════
#  SQL — Presupuesto + Hectáreas/Producción por Ccosto, parametrizado por
#  Temporada (el literal se sustituye en tiempo de ejecución, validado
#  contra un patrón de 2 dígitos — no hay input de usuario libre en la
#  consulta).
# ═══════════════════════════════════════════════════════════════════════════

SQL_PPTO_QUERY_TEMPLATE = r"""WITH FundosMapeados AS (
    SELECT FundoOriginal, FundoEstandar
    FROM (VALUES
        ('CARMEL', 'El Carmelo'),
        ('CARMELTDA', 'El Carmelo'),
        ('ELCARMELO', 'El Carmelo'),
        ('ELCARMELO2', 'El Carmelo'),
        ('CARMEN01', 'El Carmelo'),
        ('CARMEN', 'El Carmelo'),
        ('CARMELO', 'El Carmelo'),
        ('Sta Norma', 'Santa Norma'),
        ('Santa Norma', 'Santa Norma'),
        ('La Torina', 'La Torina'),
        ('La Torina Ltda', 'La Torina'),
        ('La Torina SPA', 'La Torina'),
        ('TORINASPA', 'La Torina')
    ) AS f(FundoOriginal, FundoEstandar)
),

PresuAgregado AS (
    SELECT
        LTRIM(RTRIM(CAST([Temporada] AS VARCHAR(50)))) AS Temporada,
        COALESCE(fm.FundoEstandar, LTRIM(RTRIM(CAST([Fundo] AS VARCHAR(50))))) AS Fundo,
        LTRIM(RTRIM(CAST([Ccosto] AS VARCHAR(50)))) AS CCosto,
        LTRIM(RTRIM(CAST([TipoDetG] AS VARCHAR(50)))) AS TipoDetG,
        [CCDescrip],
        [Familia]    AS CCFamilia,
        [Subfamilia] AS CCSubfamilia,
        SUM([Monto]) AS MontoPresupuesto
    FROM [Presupuesto].[dbo].[Presupuesto] p
    LEFT JOIN FundosMapeados fm ON UPPER(LTRIM(RTRIM(CAST(p.[Fundo] AS VARCHAR(50))))) = fm.FundoOriginal
    WHERE LTRIM(RTRIM(CAST([Temporada] AS VARCHAR(50)))) = '26'
    GROUP BY
        LTRIM(RTRIM(CAST([Temporada] AS VARCHAR(50)))),
        COALESCE(fm.FundoEstandar, LTRIM(RTRIM(CAST([Fundo] AS VARCHAR(50))))),
        LTRIM(RTRIM(CAST([Ccosto] AS VARCHAR(50)))),
        LTRIM(RTRIM(CAST([TipoDetG] AS VARCHAR(50)))),
        [CCDescrip], [Familia], [Subfamilia]
),

HasTotalPorCCosto AS (
    SELECT
        CAST(pf.Temporada AS VARCHAR(50)) AS Temporada,
        CAST(pf.CCosto AS VARCHAR(50)) AS CCosto,
        SUM(pf.Has) AS HasTotalProductivo,
        SUM(pf.Produccion) AS TotalCajas
    FROM [CRUZDELSUR2].[Produccion].[softland].[ProduccionFundos] pf
    WHERE pf.Especie IS NOT NULL
    AND CAST(pf.Temporada AS VARCHAR(50)) = '26'
    GROUP BY
        CAST(pf.Temporada AS VARCHAR(50)),
        CAST(pf.CCosto AS VARCHAR(50))
)

SELECT
    p.Temporada, p.Fundo, p.CCosto, p.CCFamilia, p.CCSubfamilia, p.CCDescrip, p.TipoDetG,
    p.MontoPresupuesto,
    ISNULL(hc.HasTotalProductivo, 0) AS Has,
    ISNULL(hc.TotalCajas, 0) AS Produccion
FROM PresuAgregado p
LEFT JOIN HasTotalPorCCosto hc
    ON hc.Temporada = p.Temporada
    AND hc.CCosto = p.CCosto
WHERE p.Fundo IS NOT NULL
AND LTRIM(RTRIM(p.Fundo)) <> ''
AND p.Fundo NOT IN ('Planta Almahue', 'El Comino')
AND p.CCFamilia IN ('Inversiones', 'Operaciones');"""


# Mano de Obra / Mano de Obra Cosecha: tabla dedicada (mismo servidor), ya
# trae N_Faena por fila — no requiere cruzar con Excel externo.
SQL_MANO_OBRA_QUERY_TEMPLATE = r"""WITH FundosMapeados AS (
    SELECT FundoOriginal, FundoEstandar
    FROM (VALUES
        ('CARMEL', 'El Carmelo'),
        ('CARMELTDA', 'El Carmelo'),
        ('ELCARMELO', 'El Carmelo'),
        ('ELCARMELO2', 'El Carmelo'),
        ('CARMEN01', 'El Carmelo'),
        ('CARMEN', 'El Carmelo'),
        ('CARMELO', 'El Carmelo'),
        ('Sta Norma', 'Santa Norma'),
        ('Santa Norma', 'Santa Norma'),
        ('La Torina', 'La Torina'),
        ('La Torina Ltda', 'La Torina'),
        ('La Torina SPA', 'La Torina'),
        ('TORINASPA', 'La Torina')
    ) AS f(FundoOriginal, FundoEstandar)
)
SELECT
    LTRIM(RTRIM(CAST(mo.[Temporada] AS VARCHAR(50)))) AS Temporada,
    COALESCE(fm.FundoEstandar, LTRIM(RTRIM(CAST(mo.[Fundo] AS VARCHAR(50))))) AS Fundo,
    LTRIM(RTRIM(CAST(mo.[Ccosto] AS VARCHAR(50)))) AS CCosto,
    LTRIM(RTRIM(CAST(mo.[N_Faena] AS VARCHAR(200)))) AS N_Faena,
    SUM(mo.[Monto]) AS Monto
FROM [Presupuesto].[dbo].[ManoDeObra] mo
LEFT JOIN FundosMapeados fm ON UPPER(LTRIM(RTRIM(CAST(mo.[Fundo] AS VARCHAR(50))))) = fm.FundoOriginal
WHERE LTRIM(RTRIM(CAST(mo.[Temporada] AS VARCHAR(50)))) = '26'
GROUP BY
    LTRIM(RTRIM(CAST(mo.[Temporada] AS VARCHAR(50)))),
    COALESCE(fm.FundoEstandar, LTRIM(RTRIM(CAST(mo.[Fundo] AS VARCHAR(50))))),
    LTRIM(RTRIM(CAST(mo.[Ccosto] AS VARCHAR(50)))),
    LTRIM(RTRIM(CAST(mo.[N_Faena] AS VARCHAR(200))));"""


def _sql_gastos_cfg():
    try:
        s = dict(st.secrets.get("sql_gastos", {}))
    except Exception:
        s = {}

    def pick(cands):
        for k in cands:
            for kk in (k, k.lower(), k.upper()):
                if kk in s and str(s[kk]).strip() != "":
                    return s[kk]
        return None

    server = pick(["server", "sql_server", "sql_gastos_server"])
    db     = pick(["database", "sql_database", "sql_gastos_db"])
    user   = pick(["user", "username", "sql_user", "sql_gastos_user"])
    pwd    = pick(["password", "sql_password", "sql_gastos_pwd"])
    if not all([server, db, user, pwd]):
        raise RuntimeError(
            "Faltan credenciales SQL en st.secrets['sql_gastos'] (server/database/user/password). "
            "Revisa secrets.toml.example.")
    return server, db, user, pwd


def _split_host_port(server, default_port=1433):
    s = str(server).strip()
    for sep in (",", ":"):
        if sep in s:
            host, _, port = s.partition(sep)
            try:
                return host.strip(), int(port.strip())
            except Exception:
                return host.strip(), default_port
    return s, default_port


def _tcp_reachable(host, port, timeout=5):
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except Exception:
        return False


def _sql_gastos_connect():
    server, db, user, pwd = _sql_gastos_cfg()
    host, port = _split_host_port(server)

    # Chequeo rápido de red ANTES de intentar drivers ODBC: si el host:puerto
    # no es alcanzable (ej. servidor solo accesible por VPN/red interna y
    # este proceso corre en Streamlit Cloud), fallar en segundos con un
    # mensaje claro, en vez de colgarse varios minutos probando cada
    # combinación de driver x timeout de 30s.
    if not _tcp_reachable(host, port, timeout=6):
        raise RuntimeError(
            f"No se pudo alcanzar {host}:{port} en 6 segundos. El SQL Server probablemente "
            f"solo acepta conexiones desde la red interna/VPN de la empresa, y este proceso "
            f"corre fuera de esa red (ej. Streamlit Community Cloud). Revisa la conectividad "
            f"de red antes de seguir — reintentar drivers ODBC no va a arreglar esto.")

    db_options = [db, ""]
    # Los drivers "ODBC Driver X for SQL Server" son los que hay en Windows
    # (dev local). "FreeTDS" es el que queda disponible en contenedores Linux
    # tipo Streamlit Cloud, donde no se puede instalar el driver oficial de
    # Microsoft solo con packages.txt — requiere el paquete apt tdsodbc (ver
    # packages.txt) y usa host/puerto por separado en vez de "host,puerto".
    attempts = [
        ("ODBC Driver 18 for SQL Server", f"SERVER={server}", ";TrustServerCertificate=yes;Encrypt=no"),
        ("ODBC Driver 17 for SQL Server", f"SERVER={server}", ";TrustServerCertificate=yes;Encrypt=no"),
        ("SQL Server Native Client 11.0", f"SERVER={server}", ""),
        ("SQL Server", f"SERVER={server}", ""),
        ("FreeTDS", f"SERVER={host};PORT={port}", ";TDS_Version=7.4"),
    ]
    last = None
    for drv, server_part, extra in attempts:
        for dbo in db_options:
            try:
                cs = f"DRIVER={{{drv}}};{server_part};UID={user};PWD={pwd}{extra}"
                if dbo:
                    cs += f";DATABASE={dbo}"
                return pyodbc.connect(cs, timeout=10)
            except Exception as e:
                last = e
    raise last


def _norm_temp(temp):
    temp = str(temp).strip()
    if len(temp) != 2 or not temp.isdigit():
        raise ValueError(f"Temporada inválida: {temp!r}")
    return temp


@st.cache_data(ttl=1800, show_spinner="Buscando temporadas de presupuesto disponibles...")
def load_temporadas_ppto():
    cn = _sql_gastos_connect()
    q = ("SELECT DISTINCT LTRIM(RTRIM(CAST([Temporada] AS VARCHAR(50)))) AS Temporada "
         "FROM [Presupuesto].[dbo].[Presupuesto]")
    df = pd.read_sql(q, cn)
    cn.close()
    temps = []
    for v in df["Temporada"].dropna().astype(str):
        v = v.strip()
        if v.isdigit() and len(v) == 2 and int(v) >= TEMPORADA_MIN:
            temps.append(v)
    return sorted(set(temps), reverse=True)


@st.cache_data(ttl=600, show_spinner="Consultando Presupuesto...")
def load_sql_temporada(temp):
    temp = _norm_temp(temp)
    cn = _sql_gastos_connect()
    q = SQL_PPTO_QUERY_TEMPLATE.replace("= '26'", f"= '{temp}'")
    df = pd.read_sql(q, cn)
    cn.close()
    df.columns = [str(c).strip() for c in df.columns]
    for c in ["MontoPresupuesto", "Has", "Produccion"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    for c in ["Fundo", "CCosto", "TipoDetG"]:
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip()
    return df


@st.cache_data(ttl=600, show_spinner="Consultando Mano de Obra...")
def load_mano_obra_temporada(temp):
    temp = _norm_temp(temp)
    cn = _sql_gastos_connect()
    q = SQL_MANO_OBRA_QUERY_TEMPLATE.replace("= '26'", f"= '{temp}'")
    df = pd.read_sql(q, cn)
    cn.close()
    df.columns = [str(c).strip() for c in df.columns]
    df["Monto"] = pd.to_numeric(df["Monto"], errors="coerce").fillna(0.0)
    for c in ["Fundo", "CCosto", "N_Faena"]:
        df[c] = df[c].astype(str).str.strip()
    df.loc[df["N_Faena"].isin(["", "None", "nan"]), "N_Faena"] = "(Sin Faena)"
    return df


@st.cache_data(ttl=1800, show_spinner="Buscando temporada de referencia de Hectáreas/Producción...")
def load_temporada_referencia_produccion():
    """Última Temporada con datos en ProduccionFundos (puede no coincidir con la
    Temporada de presupuesto elegida, ej. un Ppto futuro sin cosecha registrada aún)."""
    cn = _sql_gastos_connect()
    q = ("SELECT DISTINCT CAST(Temporada AS VARCHAR(50)) AS Temporada "
         "FROM [CRUZDELSUR2].[Produccion].[softland].[ProduccionFundos] WHERE Especie IS NOT NULL")
    df = pd.read_sql(q, cn)
    cn.close()
    temps = [v.strip() for v in df["Temporada"].dropna().astype(str)
             if v.strip().isdigit() and len(v.strip()) == 2]
    return max(temps) if temps else None


@st.cache_data(ttl=1800, show_spinner="Cargando Hectáreas/Producción de referencia...")
def load_has_referencia(temp_ref):
    temp_ref = _norm_temp(temp_ref)
    cn = _sql_gastos_connect()
    q = (f"SELECT CAST(CCosto AS VARCHAR(50)) AS CCosto, "
         f"SUM(Has) AS Has, SUM(Produccion) AS Produccion "
         f"FROM [CRUZDELSUR2].[Produccion].[softland].[ProduccionFundos] "
         f"WHERE Especie IS NOT NULL AND CAST(Temporada AS VARCHAR(50)) = '{temp_ref}' "
         f"GROUP BY CAST(CCosto AS VARCHAR(50))")
    df = pd.read_sql(q, cn)
    cn.close()
    df["CCosto"] = df["CCosto"].astype(str).str.strip()
    for c in ["Has", "Produccion"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    return df.set_index("CCosto")


def _norm_cc(v):
    s = str(v).strip()
    try:
        return str(int(float(s)))
    except Exception:
        return s


# ═══════════════════════════════════════════════════════════════════════════
#  MAESTRO Ccosto → Especie/Variedad (Excel de Ingresos, hoja Captura
#  Ingresos, columna Variedad2 preferida — mismo criterio que EERR
#  Proyectado del Informe de Cierre). Se usa como maestro fijo para todas
#  las temporadas de presupuesto.
# ═══════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=1800, show_spinner="Cargando maestro de Especies/Variedades...")
def load_ccosto_master():
    tok = get_token(); sid = get_site_id(tok)
    drive_id, iid = get_file_location(tok, sid, FILE_NAME)
    if iid is None:
        return pd.DataFrame(columns=["CCosto", "Especie", "Variedad"])
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{iid}/content",
        headers={"Authorization": f"Bearer {tok}"})
    r.raise_for_status()
    wb = load_workbook(BytesIO(r.content), read_only=True, data_only=True)
    sheet_name = next((s for s in SHEET_CANDS if s in wb.sheetnames), None)
    if sheet_name is None:
        wb.close()
        return pd.DataFrame(columns=["CCosto", "Especie", "Variedad"])
    ws = wb[sheet_name]
    it = ws.iter_rows(values_only=True)
    hdrs = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(next(it))]
    rows = [dict(zip(hdrs, row)) for row in it]
    wb.close()
    df = pd.DataFrame(rows)
    df.columns = [str(c).strip() for c in df.columns]

    cc_col = next((c for c in df.columns if c.strip().lower() == "ccosto"), None)
    esp_col = "Especie" if "Especie" in df.columns else None
    var_col = "Variedad2" if "Variedad2" in df.columns else ("Variedad" if "Variedad" in df.columns else None)
    if not cc_col or not esp_col:
        return pd.DataFrame(columns=["CCosto", "Especie", "Variedad"])

    out = df[[cc_col, esp_col] + ([var_col] if var_col else [])].copy()
    out.columns = ["CCosto", "Especie"] + (["Variedad"] if var_col else [])
    if "Variedad" not in out.columns:
        out["Variedad"] = "General"

    out["CCosto"] = out["CCosto"].apply(_norm_cc)
    out["Especie"] = out["Especie"].astype(str).str.strip()
    out["Variedad"] = out["Variedad"].astype(str).str.strip()
    out.loc[out["Variedad"].isin(["", "None", "nan"]), "Variedad"] = "General"
    out = out[out["CCosto"].notna() & (out["CCosto"] != "") & (out["CCosto"].str.lower() != "nan")]
    out = out[out["Especie"].notna() & (out["Especie"] != "") & (out["Especie"].str.lower() != "nan") &
              (out["Especie"].str.lower() != "general")]
    out = out.drop_duplicates(subset=["CCosto", "Especie", "Variedad"])
    return out


# ═══════════════════════════════════════════════════════════════════════════
#  AGREGACIÓN — vista Resumen (5 categorías). Especie/Variedad definen QUÉ
#  Ccostos entran (vía el maestro); Hectáreas/Producción/Costo se toman al
#  nivel del Ccosto completo (no se prorratean entre variedades de un mismo
#  Ccosto).
# ═══════════════════════════════════════════════════════════════════════════

def _ccostos_para_filtro(master, especie_sel, variedad_sel):
    m = master.copy()
    if especie_sel and especie_sel != "Todas":
        m = m[m["Especie"] == especie_sel]
    if variedad_sel and variedad_sel != "Todas":
        m = m[m["Variedad"] == variedad_sel]
    return set(m["CCosto"].unique())


def _has_por_fundo(df_raw, ccostos_sel):
    has_por_cc = df_raw.groupby(["CCosto", "Fundo"], as_index=False).agg(
        Has=("Has", "max"), Produccion=("Produccion", "max"))
    has_sel = has_por_cc[has_por_cc["CCosto"].isin(ccostos_sel)]
    out = has_sel.groupby("Fundo")[["Has", "Produccion"]].sum()
    return out.reindex(FUNDOS_ORDEN).fillna(0.0)


def _mano_obra_cosecha_split(df_labor, ccostos_norm):
    """Devuelve (mo_no_cosecha, mo_cosecha): dict Fundo -> monto."""
    mo_cosecha, mo_no_cosecha = {}, {}
    if df_labor is not None and not df_labor.empty:
        mo_sel = df_labor[df_labor["CCosto"].apply(_norm_cc).isin(ccostos_norm)]
        if not mo_sel.empty:
            cosecha_mask = mo_sel["N_Faena"].astype(str).str.lower().str.contains("cosecha", na=False)
            mo_cosecha = mo_sel[cosecha_mask].groupby("Fundo")["Monto"].sum().to_dict()
            mo_no_cosecha = mo_sel[~cosecha_mask].groupby("Fundo")["Monto"].sum().to_dict()
    return mo_no_cosecha, mo_cosecha


def build_tabla(df_raw, master, df_labor, especie_sel, variedad_sel):
    ccostos_sel = _ccostos_para_filtro(master, especie_sel, variedad_sel)
    ccostos_norm = {_norm_cc(c) for c in ccostos_sel}
    has_por_fundo = _has_por_fundo(df_raw, ccostos_sel)

    # 'Mano de obra' se excluye de la fuente Presupuesto: viene de
    # [Presupuesto].[dbo].[ManoDeObra], desglosada en Cosecha / resto.
    money_grain = df_raw.groupby(["CCosto", "Fundo", "TipoDetG"], as_index=False).agg(
        MontoPresupuesto=("MontoPresupuesto", "first"))
    money_sel = money_grain[money_grain["CCosto"].isin(ccostos_sel) &
                             (money_grain["TipoDetG"].str.lower() != "mano de obra")]
    cost_pivot = money_sel.groupby(["Fundo", "TipoDetG"], as_index=False)["MontoPresupuesto"].sum()

    def _sum_tipo(fundo, tipos_lower):
        mm = cost_pivot[(cost_pivot["Fundo"] == fundo) &
                         (cost_pivot["TipoDetG"].str.lower().isin(tipos_lower))]
        return float(mm["MontoPresupuesto"].sum())

    mo_no_cosecha, mo_cosecha = _mano_obra_cosecha_split(df_labor, ccostos_norm)

    filas = []
    for fundo in FUNDOS_ORDEN:
        has_v = float(has_por_fundo.loc[fundo, "Has"])
        prod_v = float(has_por_fundo.loc[fundo, "Produccion"])

        c_mo = float(mo_no_cosecha.get(fundo, 0.0))
        c_mo_cosecha = float(mo_cosecha.get(fundo, 0.0))
        c_fito = _sum_tipo(fundo, {"fitosanitarios"})
        c_dep = _sum_tipo(fundo, {"depreciación", "depreciacion"})
        c_gen = _sum_tipo(fundo, TIPOS_OTROS)
        c_total = c_mo + c_mo_cosecha + c_fito + c_dep + c_gen

        filas.append({
            "Fundo": fundo, "Has": has_v, "Produccion": prod_v,
            "prod_ha": (prod_v / has_v) if has_v else 0.0,
            "costo_mo": c_mo, "costo_mo_cosecha": c_mo_cosecha,
            "costo_fito": c_fito, "costo_gen": c_gen, "costo_dep": c_dep,
            "costo_total": c_total,
            "costo_mo_ha": (c_mo / has_v) if has_v else 0.0,
            "costo_mo_cosecha_ha": (c_mo_cosecha / has_v) if has_v else 0.0,
            "costo_fito_ha": (c_fito / has_v) if has_v else 0.0,
            "costo_gen_ha": (c_gen / has_v) if has_v else 0.0,
            "costo_dep_ha": (c_dep / has_v) if has_v else 0.0,
            "costo_total_ha": (c_total / has_v) if has_v else 0.0,
        })

    out = pd.DataFrame(filas)
    con_has = out["Has"] > 0
    out["rank"] = np.nan
    out.loc[con_has, "rank"] = out.loc[con_has, "costo_total_ha"].rank(method="min", ascending=True)
    return out


# ═══════════════════════════════════════════════════════════════════════════
#  AGREGACIÓN — vista Detalle (una fila por TipoDetG + Mano de Obra por
#  Faena desde [Presupuesto].[dbo].[ManoDeObra]).
# ═══════════════════════════════════════════════════════════════════════════

def build_tabla_detalle(df_raw, master, df_labor, especie_sel, variedad_sel):
    ccostos_sel = _ccostos_para_filtro(master, especie_sel, variedad_sel)
    ccostos_norm = {_norm_cc(c) for c in ccostos_sel}
    has_por_fundo = _has_por_fundo(df_raw, ccostos_sel)

    money_grain = df_raw.groupby(["CCosto", "Fundo", "TipoDetG"], as_index=False).agg(
        MontoPresupuesto=("MontoPresupuesto", "first"))
    money_sel = money_grain[money_grain["CCosto"].isin(ccostos_sel) &
                             (money_grain["TipoDetG"].str.lower() != "mano de obra")]
    cost_pivot = (money_sel.groupby(["TipoDetG", "Fundo"])["MontoPresupuesto"].sum()
                  .unstack("Fundo") if not money_sel.empty
                  else pd.DataFrame(columns=FUNDOS_ORDEN))
    cost_pivot = cost_pivot.reindex(index=TIPOS_DETALLE_ORDEN, columns=FUNDOS_ORDEN).fillna(0.0)

    faena_pivot = pd.DataFrame(columns=FUNDOS_ORDEN)
    if df_labor is not None and not df_labor.empty:
        mo_sel = df_labor[df_labor["CCosto"].apply(_norm_cc).isin(ccostos_norm)]
        if not mo_sel.empty:
            fp = (mo_sel.groupby(["N_Faena", "Fundo"])["Monto"].sum()
                  .unstack("Fundo").reindex(columns=FUNDOS_ORDEN).fillna(0.0))
            faena_pivot = fp.loc[fp.sum(axis=1).sort_values(ascending=False).index]

    return has_por_fundo, cost_pivot, faena_pivot


# ═══════════════════════════════════════════════════════════════════════════
#  FORMATO / HELPERS COMPARTIDOS
# ═══════════════════════════════════════════════════════════════════════════

def _make_formatters():
    def _f(v, d=0):
        if d == 0:
            return f"{v:,.0f}".replace(",", ".")
        return f"{v:,.{d}f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def _fh(v, d=2):
        return f"{v:,.{d}f}".replace(",", "X").replace(".", ",").replace("X", ".")

    return _f, _fh


def _rank_color(r):
    if pd.isna(r):
        return "#B0BEC5"
    if r <= 3:
        return "#388E3C"
    if r <= 7:
        return "#F9A825"
    return "#D32F2F"


def _badges_html(periodo_lbl, especie_sel, variedad_sel, extra=""):
    return f"""
<div style="margin:6px 0 14px">
  <span class="badge" style="background:rgba(84,153,199,.15);color:#5499C7;border:1px solid #5499C755">
    Selección período: {periodo_lbl} (presupuesto)
  </span>
  <span class="badge" style="background:rgba(102,187,106,.15);color:#2E7D32;border:1px solid rgba(102,187,106,.4)">
    {especie_sel} · {variedad_sel}
  </span>
  {extra}
</div>
"""


# ═══════════════════════════════════════════════════════════════════════════
#  RENDER — vista Resumen
# ═══════════════════════════════════════════════════════════════════════════

def render_resumen(df_raw, master, df_labor, periodo_lbl, especie_sel, variedad_sel):
    tabla = build_tabla(df_raw, master, df_labor, especie_sel, variedad_sel)
    _f, _fh = _make_formatters()

    st.markdown(_badges_html(periodo_lbl, especie_sel, variedad_sel), unsafe_allow_html=True)

    cols_html = "".join(f"<th>{f}</th>" for f in FUNDOS_ORDEN)

    def _row(label, key, fmt=_f):
        tds = "".join(f"<td>{fmt(r[key])}</td>" for _, r in tabla.iterrows())
        return f'<tr><td>{label}</td>{tds}</tr>'

    rows_html = ""
    rows_html += _row("Hectáreas", "Has", fmt=lambda v: _fh(v, 2))
    rows_html += _row("Producción", "Produccion", fmt=lambda v: _fh(v, 0))
    rows_html += _row("Prod x Ha", "prod_ha", fmt=lambda v: _fh(v, 1))
    rows_html += f'<tr class="sep"><td colspan="{len(FUNDOS_ORDEN)+1}">(Todos son costo x Ha)</td></tr>'
    rows_html += _row("Costo Mano de obra", "costo_mo_ha")
    rows_html += _row("Costo MO Cosecha", "costo_mo_cosecha_ha")
    rows_html += _row("Fitosanitarios", "costo_fito_ha")
    rows_html += _row("Generales", "costo_gen_ha")
    rows_html += _row("Depreciación", "costo_dep_ha")
    tot_tds = "".join(f'<td>{_f(r["costo_total_ha"])}</td>' for _, r in tabla.iterrows())
    rows_html += f'<tr class="tot"><td>Costo total x Ha</td>{tot_tds}</tr>'
    ind_tds = "".join(
        f'<td><span style="color:{_rank_color(r["rank"])}">{"-" if pd.isna(r["rank"]) else int(r["rank"])}</span></td>'
        for _, r in tabla.iterrows())
    rows_html += f'<tr class="indicador"><td>Indicador</td>{ind_tds}</tr>'

    st.markdown(
        f'<div style="overflow-x:auto;border:1px solid rgba(0,0,0,.08);border-radius:10px;'
        f'box-shadow:0 2px 10px rgba(0,0,0,.06)"><table class="tbl">'
        f'<thead><tr><th>Especie · Variedad</th>{cols_html}</tr></thead>'
        f'<tbody>{rows_html}</tbody></table></div>',
        unsafe_allow_html=True)

    st.caption("(1 el más barato, 10 el más caro — ranking solo entre fundos con hectáreas > 0 para esta especie/variedad)")

    with st.expander("🔍 Diagnóstico Mano de Obra ([Presupuesto].[dbo].[ManoDeObra]) para esta selección"):
        ccostos_sel = _ccostos_para_filtro(master, especie_sel, variedad_sel)
        ccostos_norm = {_norm_cc(c) for c in ccostos_sel}
        st.write(f"Ccostos en el filtro actual ({especie_sel} · {variedad_sel}): **{len(ccostos_sel)}**")
        if df_labor is None or df_labor.empty:
            st.warning("La tabla ManoDeObra no trajo NINGUNA fila para esta Temporada — revisa si ya se "
                       "cargó el presupuesto de mano de obra para este año.")
        else:
            mo_sel = df_labor[df_labor["CCosto"].apply(_norm_cc).isin(ccostos_norm)]
            st.write(f"Filas de ManoDeObra que matchean esos Ccostos: **{len(mo_sel)}**")
            if mo_sel.empty:
                ejemplos_mo = sorted(df_labor["CCosto"].apply(_norm_cc).unique())[:15]
                st.warning("Ningún Ccosto de ManoDeObra coincide con los de esta Especie/Variedad. "
                           f"Ccostos que SÍ existen en ManoDeObra para esta Temporada (ejemplos): {ejemplos_mo}")
            else:
                resumen_faena = (mo_sel.groupby("N_Faena")["Monto"]
                                  .agg(Monto="sum", Filas="count")
                                  .sort_values("Monto", ascending=False))
                st.write("Faenas encontradas y su monto total (todos los fundos, esta selección):")
                st.dataframe(resumen_faena, use_container_width=True)
                ccostos_con_mo = set(mo_sel["CCosto"].apply(_norm_cc).unique())
                ccostos_sin_mo = sorted(ccostos_norm - ccostos_con_mo)
                if ccostos_sin_mo:
                    st.caption(f"Ccostos del filtro SIN ninguna fila en ManoDeObra: {ccostos_sin_mo}")
                if not any("cosecha" in str(f).lower() for f in resumen_faena.index):
                    st.info("No hay ninguna faena con 'Cosecha' en el texto entre las filas encontradas — "
                            "por eso 'Costo MO Cosecha' sale en 0. Puede ser que aún no esté presupuestada "
                            "para esta temporada, o que la faena de cosecha tenga otro nombre en la tabla "
                            "(revisa arriba).")

    with st.expander("ℹ️ Metodología y supuestos de esta comparación"):
        st.markdown("""
- **Alcance actual**: solo Presupuesto (Ppto), temporadas 2024 en adelante. No se muestra "Real" por ahora.
- **Fuente de montos**: tabla `Presupuesto` (misma que usa el PBI de Gastos), filtrada por Temporada —
  **excepto Mano de Obra**, ver siguiente punto.
- **Costo Mano de obra / Costo MO Cosecha**: vienen de la tabla dedicada `[Presupuesto].[dbo].[ManoDeObra]`
  (mismo servidor SQL), agrupada por Ccosto y separada según su propia columna `N_Faena`: las filas cuyo
  N_Faena contiene "Cosecha" van a "Costo MO Cosecha", el resto a "Costo Mano de obra". Si esa tabla no
  tiene filas para la Temporada/Ccostos elegidos, ambas líneas salen en 0 (usa el diagnóstico de arriba
  para confirmar por qué).
- **Especie/Variedad por Ccosto**: viene del Excel de Ingresos (hoja "Captura Ingresos", columna `Variedad2`),
  el mismo criterio que ya usa la pestaña EERR Proyectado del Informe de Cierre — **no** se usa la columna
  `Variedad` de la tabla SQL `ProduccionFundos`, que trae códigos internos poco confiables (ej. "B062").
  Como ese Excel solo existe para la temporada actual, este mapeo Ccosto→Especie/Variedad se aplica como
  maestro fijo a todas las temporadas de presupuesto.
- **Hectáreas / Producción**: total del Ccosto completo en `ProduccionFundos` para esa temporada (no se
  prorratea si el Ccosto tiene más de una variedad).
- **Generales** agrupa: Otros Gastos, Otros Personal, Petróleo, Comercial CdS, Corrección monetaria,
  Gastos financieros.
- El costo de un Ccosto se asigna completo a cada Especie/Variedad que ahí se cultive (no se prorratea) —
  mismo criterio que ya usa el Informe de Cierre en EERR Proyectado.
""")


# ═══════════════════════════════════════════════════════════════════════════
#  RENDER — vista Detalle (oculta, ?view=detalle)
# ═══════════════════════════════════════════════════════════════════════════

def render_detalle(df_raw, master, df_labor, periodo_lbl, especie_sel, variedad_sel):
    has_por_fundo, cost_pivot, faena_pivot = build_tabla_detalle(df_raw, master, df_labor, especie_sel, variedad_sel)
    _f, _fh = _make_formatters()

    has_s = has_por_fundo["Has"]
    prod_s = has_por_fundo["Produccion"]
    prod_ha_s = (prod_s / has_s.replace(0, np.nan)).fillna(0.0)

    def _ha_div(s):
        return (s / has_s.replace(0, np.nan)).fillna(0.0)

    extra_badge = ('<span class="badge" style="background:rgba(211,47,47,.12);color:#D32F2F;'
                   'border:1px solid #D32F2F55">🔒 Vista Detalle (oculta)</span>')
    st.markdown(_badges_html(periodo_lbl, especie_sel, variedad_sel, extra=extra_badge),
                unsafe_allow_html=True)

    def _row_from_series(label, s, fmt=_f, css=""):
        tds = "".join(f"<td>{fmt(s.get(f, 0.0))}</td>" for f in FUNDOS_ORDEN)
        cls = f' class="{css}"' if css else ""
        return f'<tr{cls}><td>{label}</td>{tds}</tr>'

    mo_total_s = faena_pivot.sum(axis=0) if not faena_pivot.empty else pd.Series(0.0, index=FUNDOS_ORDEN)

    rows_html = ""
    rows_html += _row_from_series("Hectáreas", has_s, fmt=lambda v: _fh(v, 2))
    rows_html += _row_from_series("Producción", prod_s, fmt=lambda v: _fh(v, 0))
    rows_html += _row_from_series("Prod x Ha", prod_ha_s, fmt=lambda v: _fh(v, 1))
    rows_html += f'<tr class="sep"><td colspan="{len(FUNDOS_ORDEN)+1}">(Todos son costo x Ha)</td></tr>'

    rows_html += (f'<tr class="sep"><td colspan="{len(FUNDOS_ORDEN)+1}">'
                  f'MANO DE OBRA · desglose por Faena (ManoDeObra)</td></tr>')
    if faena_pivot.empty:
        rows_html += (f'<tr><td colspan="{len(FUNDOS_ORDEN)+1}" style="color:#90A4AE;font-style:italic;'
                       f'font-weight:400;text-align:left">Sin datos de Mano de Obra para esta '
                       f'Especie/Variedad/Temporada.</td></tr>')
    else:
        for faena in faena_pivot.index:
            rows_html += _row_from_series(f"· {faena}", _ha_div(faena_pivot.loc[faena]))
    rows_html += _row_from_series("Mano de Obra (Total)", _ha_div(mo_total_s), css="tot")

    rows_html += f'<tr class="sep"><td colspan="{len(FUNDOS_ORDEN)+1}">OTROS TIPODETG (Presupuesto)</td></tr>'
    for tipo in TIPOS_DETALLE_ORDEN:
        rows_html += _row_from_series(tipo, _ha_div(cost_pivot.loc[tipo]))

    total_s = _ha_div(mo_total_s + cost_pivot.sum(axis=0))
    rows_html += _row_from_series("Costo total x Ha", total_s, css="tot")

    con_has = has_s > 0
    rank = pd.Series(np.nan, index=FUNDOS_ORDEN)
    rank[con_has] = total_s[con_has].rank(method="min", ascending=True)
    ind_tds = "".join(
        f'<td><span style="color:{_rank_color(rank[f])}">{"-" if pd.isna(rank[f]) else int(rank[f])}</span></td>'
        for f in FUNDOS_ORDEN)
    rows_html += f'<tr class="indicador"><td>Indicador</td>{ind_tds}</tr>'

    cols_html = "".join(f"<th>{f}</th>" for f in FUNDOS_ORDEN)
    st.markdown(
        f'<div style="overflow-x:auto;border:1px solid rgba(0,0,0,.08);border-radius:10px;'
        f'box-shadow:0 2px 10px rgba(0,0,0,.06)"><table class="tbl">'
        f'<thead><tr><th>Especie · Variedad</th>{cols_html}</tr></thead>'
        f'<tbody>{rows_html}</tbody></table></div>',
        unsafe_allow_html=True)

    st.caption("(1 el más barato, 10 el más caro — ranking solo entre fundos con hectáreas > 0 para esta especie/variedad)")

    with st.expander("ℹ️ Metodología y supuestos — vista Detalle"):
        st.markdown("""
- Cada fila es un `TipoDetG` del presupuesto tal cual, **excepto Mano de Obra**, que en esta vista NO sale
  del presupuesto general sino de `[Presupuesto].[dbo].[ManoDeObra]`, desglosada por `N_Faena`, filtrando
  por los mismos Ccostos que ya determina el maestro de Especie/Variedad para la selección actual.
- Si esa tabla no tiene filas para la Temporada elegida, la sección de Mano de Obra sale vacía — el resto
  de los TipoDetG sigue viniendo del presupuesto igual.
- Esta vista está pensada para revisar el detalle antes de decidir si se incorpora a la vista Resumen —
  avísame qué ajustar (orden de faenas, agrupaciones, formato).
""")


# ═══════════════════════════════════════════════════════════════════════════
#  UI
# ═══════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Comparación de Presupuestos | Agua Santa", layout="wide", page_icon="🌿")
st.markdown("""
<style>
#MainMenu,footer,header{visibility:hidden}
.stApp{background:#F4F6F7}
.block-container{padding-top:1.4rem!important;max-width:100%!important}
.tbl{width:100%;border-collapse:collapse;font-size:.82rem;color:#263238}
.tbl th{background:#0C383B;color:#fff;font-size:.66rem;font-weight:700;text-transform:uppercase;
    letter-spacing:.06em;padding:8px 12px;text-align:right;white-space:nowrap}
.tbl th:first-child{text-align:left}
.tbl td{padding:6px 12px;border-bottom:1px solid rgba(0,0,0,0.06);text-align:right;
    white-space:nowrap;background:#FFFFFF;font-weight:600}
.tbl td:first-child{text-align:left;font-weight:700;color:#0C383B}
.tbl tr.sep td{background:#FFF3CD;color:#7B5800;font-weight:700;font-size:.66rem;
    text-transform:uppercase;letter-spacing:.08em;padding:5px 12px}
.tbl tr.tot td{background:#ECEFF1;font-weight:800;border-top:2px solid rgba(12,56,59,.3)}
.tbl tr.indicador td{font-weight:800}
.badge{display:inline-block;padding:4px 12px;border-radius:20px;font-size:.72rem;font-weight:700;margin-right:8px}
</style>
""", unsafe_allow_html=True)

st.markdown(f"""
<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px">
  <div>
    <div style="font-size:1.35rem;font-weight:800;color:#0C383B">🌿 Comparación de Presupuestos por Especie</div>
    <div style="font-size:.78rem;color:#546E7A">Costo x Hectárea entre Fundos — Empresas Agua Santa</div>
  </div>
  <div style="font-size:.72rem;color:#78909C">{user_email}</div>
</div>
""", unsafe_allow_html=True)

if not _PYODBC_OK:
    st.error("⚠️ pyodbc no está instalado. Ejecuta: pip install pyodbc")
    st.stop()

# Vista: 'resumen' (default, visible) o 'detalle' (oculta — solo vía ?view=detalle)
view = st.query_params.get("view", "resumen")

try:
    temporadas = load_temporadas_ppto()
except Exception as ex:
    st.error(f"⚠️ No se pudo consultar temporadas de presupuesto: {ex}")
    st.stop()

if not temporadas:
    st.warning(f"No se encontraron temporadas de presupuesto ≥ 20{TEMPORADA_MIN} en la fuente.")
    st.stop()

periodos_labels = {f"Ppto 20{t}": t for t in temporadas}

c1, c2 = st.columns([1.3, 1.7])
with c1:
    periodo_lbl = st.selectbox("📅 Período", list(periodos_labels.keys()), index=0)

temp_sel = periodos_labels[periodo_lbl]

try:
    df_raw = load_sql_temporada(temp_sel)
except Exception as ex:
    st.error(f"⚠️ No se pudo consultar la base de Presupuesto: {ex}")
    st.stop()

if df_raw.empty:
    st.warning(f"No se encontraron datos de presupuesto para {periodo_lbl} (Temporada {temp_sel}).")
    st.stop()

# ── Fallback de Hectáreas/Producción: ProduccionFundos puede no tener datos
#    todavía para la temporada de presupuesto elegida (ej. un Ppto futuro
#    recién cargado, sin producción/hectáreas registradas aún). En ese caso
#    se usan las Hectáreas/Producción de la temporada más reciente
#    disponible como referencia, dejándolo explícito en pantalla. ─────────
temp_ref_prod = None
if df_raw["Has"].sum() == 0:
    try:
        temp_ref_prod = load_temporada_referencia_produccion()
    except Exception:
        temp_ref_prod = None
    if temp_ref_prod and temp_ref_prod != temp_sel:
        try:
            has_ref = load_has_referencia(temp_ref_prod)
            df_raw["Has"] = df_raw["CCosto"].map(has_ref["Has"]).fillna(0.0)
            df_raw["Produccion"] = df_raw["CCosto"].map(has_ref["Produccion"]).fillna(0.0)
        except Exception:
            temp_ref_prod = None
    else:
        temp_ref_prod = None

try:
    master = load_ccosto_master()
except Exception as ex:
    st.error(f"⚠️ No se pudo cargar el maestro de Especies/Variedades desde SharePoint: {ex}")
    st.stop()

if master.empty:
    st.warning("El maestro de Especies/Variedades (Excel de Ingresos) vino vacío — revisa el archivo/hoja.")
    st.stop()

especies_op = sorted(master["Especie"].unique())
with c2:
    especie_sel = st.selectbox("🌿 Especie", especies_op, index=0 if especies_op else None,
                                key="especie_sel")

variedades_op = ["Todas"] + sorted(master[master["Especie"] == especie_sel]["Variedad"].unique().tolist())
variedad_sel = st.selectbox("🍇 Variedad", variedades_op, index=0, key=f"variedad_sel__{especie_sel}")

if temp_ref_prod:
    st.warning(
        f"⚠️ No hay Hectáreas/Producción registradas en ProduccionFundos para {periodo_lbl} todavía. "
        f"Se están mostrando las de la Temporada 20{temp_ref_prod} como referencia — el Costo x Ha de "
        f"abajo usa esas hectáreas, y la Producción mostrada NO es una proyección para {periodo_lbl}.")

# Mano de Obra (y Mano de Obra Cosecha) siempre viene de
# [Presupuesto].[dbo].[ManoDeObra], tanto en Resumen como en Detalle.
try:
    df_labor = load_mano_obra_temporada(temp_sel)
except Exception as ex:
    df_labor = pd.DataFrame()
    st.warning(f"⚠️ No se pudo consultar la tabla ManoDeObra: {ex}")

if view == "detalle":
    render_detalle(df_raw, master, df_labor, periodo_lbl, especie_sel, variedad_sel)
else:
    render_resumen(df_raw, master, df_labor, periodo_lbl, especie_sel, variedad_sel)

st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
if st.button("🚪 Cerrar sesión"):
    for k in ["logged_in", "login_time", "user_email"]:
        st.session_state.pop(k, None)
    st.query_params.clear()
    st.rerun()
